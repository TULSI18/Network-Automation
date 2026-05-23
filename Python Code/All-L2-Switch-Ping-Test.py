import pandas as pd
import subprocess
from pathlib import Path
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor
import openpyxl
from openpyxl.styles import PatternFill, Font

# =========================================================
# INPUT EXCEL FILE
# =========================================================

input_file = r"E:\NOC\IP Addresses\All-L2-Switch-IP.xlsx"

# =========================================================
# OUTPUT FOLDER
# =========================================================

output_folder = Path(r"E:\NOC\Output")
output_folder.mkdir(parents=True, exist_ok=True)

# =========================================================
# PING FUNCTION
# =========================================================

def ping_ip(ip):

    ip = str(ip).strip()

    # Invalid values
    if ip in ["", "-", "nan", "None", "N/A"]:
        return "Invalid IP"

    command = [
        "ping",
        "-n", "3",       # 3 packets
        "-w", "2000",    # 2 second timeout
        ip
    ]

    try:
        result = subprocess.run(
            command,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL
        )

        if result.returncode == 0:
            return "Reachable"
        else:
            return "Unreachable"

    except Exception:
        return "Error"

# =========================================================
# POP STATUS FUNCTION
# =========================================================

def get_pop_status(vlan51_status, vlan52_status):

    reachable_count = 0

    if vlan51_status == "Reachable":
        reachable_count += 1

    if vlan52_status == "Reachable":
        reachable_count += 1

    if reachable_count == 2:
        return "POP Up"

    elif reachable_count == 1:
        return "POP Partially Up"

    else:
        return "POP Down"

# =========================================================
# READ EXCEL
# =========================================================

try:

    print("\n======================================")
    print("READING INPUT FILE")
    print("======================================")

    df = pd.read_excel(input_file)

    # Clean column names
    df.columns = df.columns.str.strip()

    print("Columns Found :", df.columns.tolist())

    required_columns = [
        "District",
        "Block",
        "VLAN 51 - Jio",
        "VLAN 52-BSNL"
    ]

    # Check required columns
    for col in required_columns:
        if col not in df.columns:
            print(f"❌ Missing Column : {col}")
            exit()

    # =====================================================
    # CREATE TASK LIST
    # =====================================================

    tasks = []

    for index, row in df.iterrows():

        tasks.append(("VLAN51", index, row["VLAN 51 - Jio"]))
        tasks.append(("VLAN52", index, row["VLAN 52-BSNL"]))

    # =====================================================
    # MULTITHREAD PING
    # =====================================================

    print("\n======================================")
    print("PING STARTED")
    print("======================================")

    results = {}

    def worker(task):

        vlan_type, index, ip = task

        status = ping_ip(ip)

        print(f"{ip} --> {status}")

        return (vlan_type, index, status)

    with ThreadPoolExecutor(max_workers=75) as executor:

        for vlan_type, index, status in executor.map(worker, tasks):

            if index not in results:
                results[index] = {}

            results[index][vlan_type] = status

    # =====================================================
    # ADD STATUS COLUMNS
    # =====================================================

    vlan51_status_list = []
    vlan52_status_list = []
    pop_status_list = []

    for index in df.index:

        vlan51_status = results[index].get("VLAN51", "Error")
        vlan52_status = results[index].get("VLAN52", "Error")

        pop_status = get_pop_status(vlan51_status, vlan52_status)

        vlan51_status_list.append(vlan51_status)
        vlan52_status_list.append(vlan52_status)
        pop_status_list.append(pop_status)

    df["VLAN51 Status"] = vlan51_status_list
    df["VLAN52 Status"] = vlan52_status_list
    df["POP Status"] = pop_status_list

    # =====================================================
    # SUMMARY
    # =====================================================

    total_pop = len(df)

    pop_up = len(df[df["POP Status"] == "POP Up"])
    partial_up = len(df[df["POP Status"] == "POP Partially Up"])
    pop_down = len(df[df["POP Status"] == "POP Down"])

    print("\n======================================")
    print("SUMMARY")
    print("======================================")

    print(f"Total POP           : {total_pop}")
    print(f"POP Up              : {pop_up}")
    print(f"POP Partially Up    : {partial_up}")
    print(f"POP Down            : {pop_down}")

    # =====================================================
    # OUTPUT FILE
    # =====================================================

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    output_file = output_folder / f"L2_Switch_Ping_Report_{timestamp}.xlsx"

    # Save excel
    df.to_excel(output_file, index=False)

    # =====================================================
    # COLOR FORMATTING
    # =====================================================

    wb = openpyxl.load_workbook(output_file)
    ws = wb.active

    # Header formatting
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Status colors
    green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
    red_fill = PatternFill(start_color="FF6666", end_color="FF6666", fill_type="solid")

    # Find POP Status column
    pop_status_col = None

    for cell in ws[1]:
        if cell.value == "POP Status":
            pop_status_col = cell.column
            break

    # Apply colors
    for row in range(2, ws.max_row + 1):

        cell = ws.cell(row=row, column=pop_status_col)

        if cell.value == "POP Up":
            cell.fill = green_fill

        elif cell.value == "POP Partially Up":
            cell.fill = yellow_fill

        elif cell.value == "POP Down":
            cell.fill = red_fill

    # Auto column width
    for column_cells in ws.columns:

        length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)

        ws.column_dimensions[column_cells[0].column_letter].width = length + 5

    wb.save(output_file)

    # =====================================================
    # COMPLETED
    # =====================================================

    print("\n======================================")
    print("✅ PING CHECK COMPLETED")
    print("======================================")

    print(f"\n📁 Output File : {output_file}")

except Exception as e:

    print("\n❌ SCRIPT FAILED")
    print("Error :", e)