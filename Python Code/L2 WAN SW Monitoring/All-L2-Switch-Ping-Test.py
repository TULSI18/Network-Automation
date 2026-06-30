import pandas as pd
import subprocess
import re
import time
from pathlib import Path
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

# =========================================================
# START TIMER
# =========================================================

start_time = time.time()

# =========================================================
# INPUT EXCEL FILE
# =========================================================

input_file = r"E:\NOC\IP Addresses\All-L2-Switch-IP.xlsx"

# =========================================================
# OUTPUT FOLDER
# =========================================================

output_folder = Path(r"E:\NOC\Output\L2-WAN-SW-Monitoring")
output_folder.mkdir(parents=True, exist_ok=True)

# =========================================================
# PING FUNCTION
# =========================================================

def ping_ip(ip):

    ip = str(ip).strip()

    # Invalid values
    if ip in ["", "-", "nan", "None", "N/A"]:
        return {
            "status": "Invalid IP",
            "rtt": "N/A",
            "loss": "100%"
        }

    command = [
        "ping",
        "-n", "3",        # 3 packets
        "-w", "1000",     # 1 second timeout
        ip
    ]

    try:

        result = subprocess.run(
            command,
            capture_output=True,
            text=True
        )

        output = result.stdout

        # =================================================
        # PACKET LOSS
        # =================================================

        loss_match = re.search(r"\((\d+)% loss\)", output)

        if loss_match:
            loss = f"{loss_match.group(1)}%"
        else:
            loss = "100%"

        # =================================================
        # RTT EXTRACTION
        # =================================================

        rtt_match = re.search(r"Average = (\d+)ms", output)

        if rtt_match:
            rtt = f"{rtt_match.group(1)}ms"
        else:
            rtt = "N/A"

        # =================================================
        # STATUS
        # =================================================

        if result.returncode == 0:
            status = "Reachable"
        else:
            status = "Unreachable"

        print(f"{ip} --> {status} | RTT={rtt} | Loss={loss}")

        return {
            "status": status,
            "rtt": rtt,
            "loss": loss
        }

    except Exception:

        return {
            "status": "Error",
            "rtt": "N/A",
            "loss": "100%"
        }

# =========================================================
# POP STATUS FUNCTION
# =========================================================

def get_pop_status(vlan51_status, vlan52_status):

    reachable_count = 0

    if vlan51_status == "Reachable":
        reachable_count += 1

    if vlan52_status == "Reachable":
        reachable_count += 1

    if reachable_count == 2:
        return "POP Up"

    elif reachable_count == 1:
        return "POP Partially Up"

    else:
        return "POP Down"

# =========================================================
# READ EXCEL
# =========================================================

try:

    print("\n======================================")
    print("READING INPUT FILE")
    print("======================================")

    df = pd.read_excel(input_file)

    # Clean column names
    df.columns = df.columns.str.strip()

    print("Columns Found :", df.columns.tolist())

    required_columns = [
        "District",
        "Block",
        "VLAN 51 - Jio",
        "VLAN 52-BSNL"
    ]

    # =====================================================
    # CHECK REQUIRED COLUMNS
    # =====================================================

    for col in required_columns:
        if col not in df.columns:
            print(f"❌ Missing Column : {col}")
            exit()

    # =====================================================
    # CREATE TASKS
    # =====================================================

    tasks = []

    for index, row in df.iterrows():

        tasks.append(("VLAN51", index, row["VLAN 51 - Jio"]))
        tasks.append(("VLAN52", index, row["VLAN 52-BSNL"]))

    # =====================================================
    # MULTITHREAD PING
    # =====================================================

    print("\n======================================")
    print("PING STARTED")
    print("======================================")

    results = {}

    def worker(task):

        vlan_type, index, ip = task

        result = ping_ip(ip)

        return (vlan_type, index, result)

    with ThreadPoolExecutor(max_workers=75) as executor:

        for vlan_type, index, result in executor.map(worker, tasks):

            if index not in results:
                results[index] = {}

            results[index][vlan_type] = result

    # =====================================================
    # CREATE OUTPUT COLUMNS
    # =====================================================

    vlan51_status_list = []
    vlan51_rtt_list = []
    vlan51_loss_list = []

    vlan52_status_list = []
    vlan52_rtt_list = []
    vlan52_loss_list = []

    pop_status_list = []

    for index in df.index:

        vlan51 = results[index].get("VLAN51", {})
        vlan52 = results[index].get("VLAN52", {})

        vlan51_status = vlan51.get("status", "Error")
        vlan51_rtt = vlan51.get("rtt", "N/A")
        vlan51_loss = vlan51.get("loss", "100%")

        vlan52_status = vlan52.get("status", "Error")
        vlan52_rtt = vlan52.get("rtt", "N/A")
        vlan52_loss = vlan52.get("loss", "100%")

        pop_status = get_pop_status(vlan51_status, vlan52_status)

        vlan51_status_list.append(vlan51_status)
        vlan51_rtt_list.append(vlan51_rtt)
        vlan51_loss_list.append(vlan51_loss)

        vlan52_status_list.append(vlan52_status)
        vlan52_rtt_list.append(vlan52_rtt)
        vlan52_loss_list.append(vlan52_loss)

        pop_status_list.append(pop_status)

    # =====================================================
    # ADD COLUMNS TO DATAFRAME
    # =====================================================

    df["VLAN51 Status"] = vlan51_status_list
    df["VLAN51 RTT"] = vlan51_rtt_list
    df["VLAN51 Loss"] = vlan51_loss_list

    df["VLAN52 Status"] = vlan52_status_list
    df["VLAN52 RTT"] = vlan52_rtt_list
    df["VLAN52 Loss"] = vlan52_loss_list

    df["POP Status"] = pop_status_list

    # =====================================================
    # DOWN POP SHEET
    # =====================================================

    down_df = df[
        (df["POP Status"] == "POP Down") |
        (df["POP Status"] == "POP Partially Up")
    ]

    # =====================================================
    # SUMMARY SHEET
    # =====================================================

    total_pop = len(df)

    pop_up = len(df[df["POP Status"] == "POP Up"])

    partial_up = len(df[df["POP Status"] == "POP Partially Up"])

    pop_down = len(df[df["POP Status"] == "POP Down"])

    summary_data = {
        "Metric": [
            "Total POP",
            "POP Up",
            "POP Partially Up",
            "POP Down"
        ],
        "Count": [
            total_pop,
            pop_up,
            partial_up,
            pop_down
        ]
    }

    summary_df = pd.DataFrame(summary_data)

    # =====================================================
    # OUTPUT FILE
    # =====================================================

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    output_file = output_folder / f"POP_Status_Report_{timestamp}.xlsx"

    # =====================================================
    # WRITE MULTIPLE SHEETS
    # =====================================================

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:

        df.to_excel(writer, sheet_name="POP_Status", index=False)

        down_df.to_excel(writer, sheet_name="Down_POPs", index=False)

        summary_df.to_excel(writer, sheet_name="Summary", index=False)

    # =====================================================
    # FORMAT EXCEL
    # =====================================================

    wb = openpyxl.load_workbook(output_file)

    # =====================================================
    # COLORS
    # =====================================================

    green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")

    yellow_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")

    red_fill = PatternFill(start_color="FF6666", end_color="FF6666", fill_type="solid")

    blue_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

    white_font = Font(color="FFFFFF", bold=True)

    center_align = Alignment(horizontal="center", vertical="center")

    # =====================================================
    # FORMAT ALL SHEETS
    # =====================================================

    for ws in wb.worksheets:

        # Header Formatting
        for cell in ws[1]:

            cell.fill = blue_fill
            cell.font = white_font
            cell.alignment = center_align

        # Auto Width
        for column_cells in ws.columns:

            length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)

            ws.column_dimensions[column_cells[0].column_letter].width = length + 5

        # Freeze Header
        ws.freeze_panes = "A2"

    # =====================================================
    # FORMAT POP STATUS SHEET
    # =====================================================

    ws = wb["POP_Status"]

    # Find columns
    headers = {}

    for cell in ws[1]:
        headers[cell.value] = cell.column

    for row in range(2, ws.max_row + 1):

        # POP STATUS
        pop_cell = ws.cell(row=row, column=headers["POP Status"])

        if pop_cell.value == "POP Up":
            pop_cell.fill = green_fill

        elif pop_cell.value == "POP Partially Up":
            pop_cell.fill = yellow_fill

        elif pop_cell.value == "POP Down":
            pop_cell.fill = red_fill

        # VLAN STATUS COLORS
        for status_col in ["VLAN51 Status", "VLAN52 Status"]:

            cell = ws.cell(row=row, column=headers[status_col])

            if cell.value == "Reachable":
                cell.fill = green_fill

            elif cell.value == "Unreachable":
                cell.fill = red_fill

            elif cell.value == "Invalid IP":
                cell.fill = yellow_fill

    # =====================================================
    # SAVE WORKBOOK
    # =====================================================

    wb.save(output_file)

    # =====================================================
    # EXECUTION TIME
    # =====================================================

    end_time = time.time()

    execution_time = round(end_time - start_time, 2)

    # =====================================================
    # FINAL OUTPUT
    # =====================================================

    print("\n======================================")
    print("✅ PING CHECK COMPLETED")
    print("======================================")

    print(f"\n📁 Output File : {output_file}")

    print(f"\n⏱ Execution Time : {execution_time} seconds")

    print("\n======================================")
    print("SUMMARY")
    print("======================================")

    print(f"Total POP           : {total_pop}")
    print(f"POP Up              : {pop_up}")
    print(f"POP Partially Up    : {partial_up}")
    print(f"POP Down            : {pop_down}")

except Exception as e:

    print("\n❌ SCRIPT FAILED")
    print("Error :", e)